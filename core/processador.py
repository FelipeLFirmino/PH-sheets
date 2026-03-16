import re
import json
import pandas as pd
import xml.etree.ElementTree as ET
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ─── Padrão embalagem ────────────────────────────────────────────────────────
_PATTERN_EMB = re.compile(
    r'(?:^(\d+)\s|(?:CAIXA COM|PACOTE COM|KIT COM|PCT\s*C/|CX\s*C/|C/)\s*(\d+))',
    re.IGNORECASE
)

# ─── Mapeamento de colunas (1-based) ─────────────────────────────────────────
#
# FLUXO VAREJO:
#   1. C_REAL  = NF_U × MULT
#   2. FRETE   = C_REAL × FRETE%
#   3. DESP    = C_REAL × DESP%
#   4. CRED    = NF_U × CRED%  (zerado se ST/CST isento — ANT não zera)
#   5. C_ENT   = C_REAL + ST + ANT + IPI + FRETE + DESP − CRED
#   6. FEDERAL = P_VAR × FED%
#   7. CARTÃO  = P_VAR × CART%
#   8. ICMS_S  = P_VAR × ICMS%
#   9. C_SAIDA = C_ENT + FEDERAL + CARTÃO + ICMS_S
#  10. P_MIN   = C_SAIDA ÷ (1 − META%)
#
# FLUXO ATACADO:
#  11. NF_ATC     = NF_U × MULT_ATC          (parâm col 25 linha 2)
#  12. P_ATC      = P_VAR × (1 − DESC_ATC%)  (parâm col 26 linha 2)
#  13. FED_ATC    = NF_ATC × FED%
#  14. CART_ATC   = P_ATC × CART%
#  15. ICM_ATC    = NF_ATC × ICM%
#  16. C_SAIDA_ATC= C_ENT + FED_ATC + CART_ATC + ICM_ATC
#  17. MARGEM_ATC = (P_ATC − C_SAIDA_ATC) / P_ATC
#
# A  B     C    D    E    F      G     H      I      J      K      L
# NF DESC  REF  SKU  QTD  NF_U   ST_U  ANT_U  IPI_U  C_REAL FRETE  DESP
#  1  2     3    4    5    6      7     8      9      10     11     12
#
# M      N    O      P      Q      R       S       T     U      V       W      X
# CRED   CST  C_ENT  FED    CART   ICMS_S  C_SAIDA META  P_MIN  P_ATUAL P_VAR  MARGEM
# 13     14   15     16     17     18      19      20    21     22      23     24
#
# Y       Z      AA       AB        AC      AD           AE          AF         AG
# NF_ATC  P_ATC  FED_ATC  CART_ATC  ICM_ATC C_SAIDA_ATC  MARGEM_ATC  AUDIT_SYS  AUDIT_EMB
# 25      26     27       28        29      30           31          32         33

COL = {
    'NF': 1, 'DESC': 2, 'REF': 3, 'SKU': 4, 'QTD': 5,
    'NF_U': 6, 'ST_U': 7, 'ANT_U': 8, 'IPI_U': 9,
    'C_REAL': 10, 'FRETE': 11, 'DESP': 12, 'CRED': 13,
    'CST': 14, 'C_ENT': 15,
    'FED': 16, 'CARTAO': 17, 'ICMS_S': 18, 'C_SAIDA': 19,
    'META': 20, 'P_MIN': 21,
    'P_ATUAL': 22, 'P_VAR': 23, 'MARGEM': 24,
    'NF_ATC': 25, 'P_ATC': 26,
    'FED_ATC': 27, 'CART_ATC': 28, 'ICM_ATC': 29,
    'C_SAIDA_ATC': 30, 'MARGEM_ATC': 31,
    'AUDIT_SYS': 32, 'AUDIT_EMB': 33,
}
TOTAL_COLS = 33


# ─── Helpers ─────────────────────────────────────────────────────────────────
def limpar_str(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    s = str(val).replace('="', '').replace('"', '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def limpar_preco(val):
    if pd.isna(val) or str(val).strip() == "":
        return 0.0
    try:
        return float(str(val).replace('R$', '').replace('.', '').replace(',', '.').strip())
    except:
        return 0.0


def get_xml_text(node, xpath, ns, default=""):
    if node is None:
        return default
    child = node.find(xpath, ns)
    if child is not None and child.text is not None:
        return child.text.strip()
    return default


def extrair_qtd_embalagem(desc_xml, v_un_xml, p_sys, mult):
    match = _PATTERN_EMB.search(desc_xml)
    if not match:
        return 1
    qtd = int(match.group(1))
    if qtd <= 1:
        return 1
    if p_sys > 0 and (p_sys * qtd) > (v_un_xml * mult * 3):
        return 1
    return qtd


# ─── Extração XML + merge CSV ─────────────────────────────────────────────────
def gerar_tabela(xml_path, csv_path, fornecedor, nota_ref, params):
    try:
        P_MULT     = float(params.get('mult_var',  2.0))
        P_DESP     = float(params.get('desp',      10))   / 100
        P_FRETE    = float(params.get('frete',     10))   / 100
        P_CRED     = float(params.get('cred_icms', 4))    / 100
        P_FED      = float(params.get('fed',       9.13)) / 100
        P_ICM      = float(params.get('icm',       21))   / 100
        P_CART     = float(params.get('cartao',    4))    / 100
        P_MULT_ATC = float(params.get('mult_atc',  1.3))
        P_DESC_ATC = float(params.get('desc_atc',  15))   / 100

        tree = ET.parse(xml_path)
        ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
        infNFe    = tree.getroot().find('.//nfe:infNFe', ns)
        num_nf    = get_xml_text(infNFe, './/nfe:ide/nfe:nNF', ns, "000")
        chave_nfe = infNFe.attrib.get('Id', '')[3:] if infNFe is not None else ""

        # API SEFAZ AL
        dados_api = []
        try:
            url = (
                "https://contribuinte.sefaz.al.gov.br/cobrancadfe/"
                f"sfz-cobranca-dfe-api/api/detalhe-calculo-nfes?chaveNota.equals={chave_nfe}"
            )
            res = requests.get(
                url,
                headers={'Accept': 'application/json', 'User-Agent': 'Mozilla/5.0'},
                timeout=10
            )
            print(f"\n{'='*70}")
            print(f"[SEFAZ API] chave: {chave_nfe}")
            print(f"[SEFAZ API] status: {res.status_code}")
            if res.status_code == 200:
                dados_api = res.json()
                print(f"[SEFAZ API] {len(dados_api)} item(s) retornados:")
                for idx, it in enumerate(dados_api):
                    print(f"  [{idx}] tipo={it.get('tipoImposto')!r:6}  "
                          f"icms={it.get('valorIcmsCalculado')!r}  "
                          f"fecoep={it.get('valorFecoepCalculado')!r}  "
                          f"desc={str(it.get('descricaoProduto',''))[:60]!r}")
            else:
                print(f"[SEFAZ API] resposta inesperada: {res.text[:200]}")
            print(f"{'='*70}\n")
        except Exception as _api_err:
            print(f"[SEFAZ API] ERRO na requisição: {_api_err}")

        itens_xml = []
        for det in tree.getroot().findall('.//nfe:det', ns):
            p = det.find('nfe:prod', ns)
            i = det.find('nfe:imposto', ns)

            desc_xml = get_xml_text(p, 'nfe:xProd', ns, "SEM DESCRICAO")
            v_st_xml = float(get_xml_text(i, './/nfe:vICMSST', ns, "0"))

            v_st_api, v_ant_api = 0.0, 0.0
            matches_log = []
            for item in dados_api:
                api_desc = str(item.get('descricaoProduto', '')).strip().upper()
                if api_desc in desc_xml.strip().upper():
                    vt = (float(item.get('valorIcmsCalculado', 0) or 0)
                          + float(item.get('valorFecoepCalculado', 0) or 0))
                    tipo = item.get('tipoImposto')
                    if tipo == 'ANT':
                        v_ant_api += vt
                    else:
                        v_st_api += vt
                    matches_log.append(f"tipo={tipo!r} valor={vt:.4f} api_desc={api_desc[:40]!r}")

            print(f"[PRODUTO] {desc_xml[:60]!r}")
            print(f"  vICMSST (XML) = {v_st_xml:.4f}")
            if matches_log:
                print(f"  matches SEFAZ API ({len(matches_log)}):")
                for ml in matches_log:
                    print(f"    → {ml}")
            else:
                print(f"  matches SEFAZ API: nenhum")
            print(f"  v_st_api={v_st_api:.4f}  v_ant_api={v_ant_api:.4f}  "
                  f"vST_total={v_st_xml + v_st_api:.4f}")

            cst = ""
            icms_node = i.find('.//nfe:ICMS', ns) if i is not None else None
            if icms_node is not None:
                for child in icms_node:
                    tag = child.find('nfe:CST', ns)
                    if tag is None:
                        tag = child.find('nfe:CSOSN', ns)
                    if tag is not None and tag.text is not None:
                        cst = tag.text.strip()
                        break

            ean_xml = limpar_str(get_xml_text(p, 'nfe:cEAN', ns))
            if not ean_xml or ean_xml.upper() == "SEM GTIN":
                ean_xml = "SEM GTIN"

            itens_xml.append({
                'ean_xml':  ean_xml,
                'ref_xml':  limpar_str(get_xml_text(p, 'nfe:cProd', ns)),
                'desc_xml': desc_xml,
                'qCom':     float(get_xml_text(p, 'nfe:qCom',   ns, "1")),
                'vUnCom':   float(get_xml_text(p, 'nfe:vUnCom', ns, "0")),
                'vProd':    float(get_xml_text(p, 'nfe:vProd',  ns, "0")),
                'vIPI':     float(get_xml_text(i, './/nfe:vIPI', ns, "0")),
                'vST':      v_st_xml + v_st_api,
                'vANT':     v_ant_api,
                'nf_base':  num_nf,
                'cst':      cst,
            })

        df_xml = pd.DataFrame(itens_xml)

        # CSV do sistema — FutureWarning corrigido: atribuição via .loc[]
        try:
            df_sys = pd.read_csv(csv_path, sep=';', encoding='utf-8',  on_bad_lines='skip', dtype=str)
        except:
            df_sys = pd.read_csv(csv_path, sep=';', encoding='latin1', on_bad_lines='skip', dtype=str)

        df_sys.columns = df_sys.columns.str.replace('"', '').str.strip().str.upper()
        df_sys = df_sys.copy()  # evita SettingWithCopyWarning
        df_sys.loc[:, 'ean_sys']   = df_sys['BARRA'].apply(limpar_str)      if 'BARRA'      in df_sys.columns else ''
        df_sys.loc[:, 'ref_sys']   = df_sys['REFERÊNCIA'].apply(limpar_str) if 'REFERÊNCIA' in df_sys.columns else ''
        df_sys.loc[:, 'preco_sys'] = df_sys['PREÇO'].apply(limpar_preco)    if 'PREÇO'      in df_sys.columns else 0.0

        # Merge por EAN
        df_base = pd.merge(df_xml, df_sys[['ean_sys', 'ref_sys', 'preco_sys']],
                           left_on='ean_xml', right_on='ean_sys', how='left')

        # Fallback por REF
        mask = (df_base['preco_sys'].isna()) | (df_base['preco_sys'] == 0)
        if mask.any():
            uniq  = df_sys.drop_duplicates(subset=['ref_sys']).dropna(subset=['ref_sys'])
            mp    = uniq.set_index('ref_sys')['preco_sys'].to_dict()
            me    = uniq.set_index('ref_sys')['ean_sys'].to_dict()
            df_base.loc[mask, 'preco_sys'] = df_base.loc[mask, 'ref_xml'].map(mp)
            df_base.loc[mask, 'ean_sys']   = df_base.loc[mask, 'ref_xml'].map(me)

        df_base = df_base.copy()
        df_base.loc[:, 'preco_sys'] = df_base['preco_sys'].fillna(0.0)
        df_base.loc[:, 'sku_final'] = df_base.apply(
            lambda r: r['ean_sys'] if pd.notna(r['ean_sys']) and str(r['ean_sys']).strip() != ""
                      else r['ean_xml'], axis=1)
        df_base.loc[:, 'qtd_emb']   = df_base.apply(
            lambda r: extrair_qtd_embalagem(r['desc_xml'], r['vUnCom'], r['preco_sys'], P_MULT),
            axis=1)

        rows = []
        for _, row in df_base.iterrows():
            q       = max(float(row['qCom']), 1)
            nf_u    = round(row['vProd'] / q, 2)
            st_u    = round(row['vST']   / q, 2)
            ant_u   = round(row['vANT']  / q, 2)
            ipi_u   = round(row['vIPI']  / q, 2)
            cst     = str(row['cst'])
            qtd_emb = int(row['qtd_emb'])

            p_sys_val        = float(row['preco_sys']) if pd.notna(row['preco_sys']) else 0.0
            preco_venda_base = round(p_sys_val * qtd_emb, 2) if p_sys_val > 0.01 else 0.0

            rows.append({
                'nf':        row['nf_base'],
                'desc':      row['desc_xml'],
                'ref':       row['ref_xml'],
                'sku':       row['sku_final'],
                'qtd':       q,
                'nf_u':      nf_u,
                'st_u':      st_u,
                'ant_u':     ant_u,
                'ipi_u':     ipi_u,
                'cst':       cst,
                'p_atual':   preco_venda_base,
                'p_sys_raw': p_sys_val,
                'qtd_emb':   qtd_emb,
            })

        params_out = {
            'mult': P_MULT, 'frete': P_FRETE, 'desp': P_DESP,
            'cred': P_CRED, 'fed': P_FED, 'icm': P_ICM, 'cartao': P_CART,
            'mult_atc': P_MULT_ATC, 'desc_atc': P_DESC_ATC,
        }
        return True, (rows, params_out, num_nf)

    except Exception as e:
        import traceback
        return False, traceback.format_exc()


# ─── Estilos ─────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)

def _borda():
    s = Side("thin")
    return Border(left=s, right=s, top=s, bottom=s)

F_AZUL     = _fill("BDD7EE")
F_AMAR     = _fill("FFF2CC")
F_AMAR_ANT = _fill("D1FAE5")
F_PELE     = _fill("FCE4D6")
F_CINZA    = _fill("D9D9D9")
F_VERDE    = _fill("E2EFDA")
F_PARAM    = _fill("F2F2F2")
BORDA   = _borda()
ALI_CTR = Alignment(horizontal='center', vertical='center')


def _c(ws, r, c, value=None, fill=None, font=None, fmt=None, bold=False):
    """Atalho para setar célula com valor + estilo."""
    cell = ws.cell(row=r, column=c, value=value)
    cell.border    = BORDA
    cell.alignment = ALI_CTR
    if fill: cell.fill          = fill
    if fmt:  cell.number_format = fmt
    if bold: cell.font          = Font(bold=True)
    if font: cell.font          = font
    return cell


# ─── Salvar Excel ─────────────────────────────────────────────────────────────
def salvar_excel_estilizado(dados, path):
    rows, P, num_nf = dados
    L = get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Precificação'

    # ── Linha 1: Cabeçalhos ──────────────────────────────────────────────────
    headers = [
        'NF', 'DESCRIÇÃO', 'REF', 'SKU', 'QTD',
        'NF UNIT', 'ST UNIT', 'ANT UNIT', 'IPI UNIT',
        'CUSTO REAL', 'FRETE', 'DESPESA', 'CRED. ICMS',
        'CST', 'CUSTO ENTRADA',
        'FEDERAL', 'CARTÃO', 'ICMS SAÍDA', 'CUSTO SAÍDA',
        'META %', 'PREÇO MÍN VIÁVEL',
        'PREÇO ATUAL', 'PREÇO VAREJO', 'MARGEM REAL',
        'NF ATC', 'PREÇO ATC',
        'FEDERAL ATC', 'CARTÃO ATC', 'ICMS ATC',
        'CUSTO SAÍDA ATC', 'MARGEM ATC',
        'P.UNIT SISTEMA', 'QTD EMB',
    ]
    for c, h in enumerate(headers, 1):
        _c(ws, 1, c, h, fill=F_PARAM, bold=True)

    # ── Linha 2: Parâmetros ──────────────────────────────────────────────────
    # col 10 = MULT (número), cols % = formatados como percentual
    param_vals = {
        COL['C_REAL']:  P['mult'],       # col 10 — multiplicador varejo
        COL['FRETE']:   P['frete'],      # col 11
        COL['DESP']:    P['desp'],       # col 12
        COL['CRED']:    -P['cred'],      # col 13 — negativo
        COL['FED']:     P['fed'],        # col 16
        COL['CARTAO']:  P['cartao'],     # col 17
        COL['ICMS_S']:  P['icm'],        # col 18
        COL['NF_ATC']:  P['mult_atc'],   # col 25 — multiplicador atacado
        COL['P_ATC']:   P['desc_atc'],   # col 26 — desconto atacado
    }
    pct_cols = {COL['FRETE'], COL['DESP'], COL['CRED'], COL['FED'], COL['CARTAO'], COL['ICMS_S'], COL['P_ATC']}
    for c in range(1, TOTAL_COLS + 1):
        val = param_vals.get(c)
        fmt = '0.00%' if c in pct_cols else ('0.00' if c == COL['C_REAL'] else 'General')
        _c(ws, 2, c, val, fill=F_PARAM, fmt=fmt)
    ws.cell(2, COL['META']).value = "META %"

    # ── Linhas de dados (linha 3 em diante) ──────────────────────────────────
    for idx, row in enumerate(rows):
        r       = idx + 3
        has_st  = row['st_u']  > 0.005
        has_ant = row['ant_u'] > 0.005 and row['st_u'] <= 0.005
        cst     = row['cst']

        # Letras das colunas para fórmulas
        F  = L(COL['NF_U'])       # NF UNIT
        G  = L(COL['ST_U'])       # ST
        H  = L(COL['ANT_U'])      # ANT
        I  = L(COL['IPI_U'])      # IPI
        N  = L(COL['C_REAL'])     # CUSTO REAL
        J  = L(COL['FRETE'])      # FRETE
        K  = L(COL['DESP'])       # DESPESA
        Lc = L(COL['CRED'])       # CRED ICMS
        Oc = L(COL['CST'])        # CST
        M  = L(COL['C_ENT'])      # CUSTO ENTRADA
        P2 = L(COL['FED'])        # FEDERAL
        Qc = L(COL['CARTAO'])     # CARTÃO
        Rc = L(COL['ICMS_S'])     # ICMS SAÍDA
        S  = L(COL['C_SAIDA'])    # CUSTO SAÍDA
        T  = L(COL['META'])       # META %
        V  = L(COL['P_ATUAL'])    # PREÇO ATUAL
        W  = L(COL['P_VAR'])      # PREÇO VAREJO
        Ya = L(COL['NF_ATC'])     # NF ATC
        Za = L(COL['P_ATC'])      # PREÇO ATC
        AA = L(COL['FED_ATC'])    # FEDERAL ATC
        AB = L(COL['CART_ATC'])   # CARTÃO ATC
        AC = L(COL['ICM_ATC'])    # ICMS ATC
        AD = L(COL['C_SAIDA_ATC'])# CUSTO SAÍDA ATC
        AE = L(COL['MARGEM_ATC']) # MARGEM ATC

        # Valores fixos (vêm do XML/CSV)
        _c(ws, r, COL['NF'],    row['nf'])
        _c(ws, r, COL['DESC'],  row['desc'])
        _c(ws, r, COL['REF'],   row['ref'])
        _c(ws, r, COL['SKU'],   row['sku'])
        _c(ws, r, COL['QTD'],   row['qtd'])
        _c(ws, r, COL['NF_U'],  row['nf_u'],  fmt='#,##0.00')
        _c(ws, r, COL['ST_U'],  row['st_u'],  fmt='#,##0.00')
        _c(ws, r, COL['ANT_U'], row['ant_u'], fmt='#,##0.00')
        _c(ws, r, COL['IPI_U'], row['ipi_u'], fmt='#,##0.00')
        _c(ws, r, COL['CST'],   cst)
        _c(ws, r, COL['P_ATUAL'], row['p_atual'], fmt='#,##0.00')

        # META % — editável por produto, padrão 15%
        mc = ws.cell(row=r, column=COL['META'], value=0.15)
        mc.border = BORDA; mc.alignment = ALI_CTR
        mc.number_format = '0.00%'
        mc.fill = F_VERDE; mc.font = Font(bold=True)

        # Colunas de auditoria
        _c(ws, r, COL['AUDIT_SYS'], row['p_sys_raw'], fill=F_CINZA,
           font=Font(italic=True, color="888888"), fmt='#,##0.00')
        _c(ws, r, COL['AUDIT_EMB'], row['qtd_emb'],   fill=F_CINZA,
           font=Font(italic=True, color="888888"))

        # ── Fórmulas ─────────────────────────────────────────────────────────

        # N - CUSTO REAL = NF_U × MULT  ← multiplicador aplicado PRIMEIRO
        ws.cell(r, COL['C_REAL']).value = f"=ROUND({F}{r}*{N}$2,2)"

        # J - FRETE = CUSTO REAL × FRETE%
        ws.cell(r, COL['FRETE']).value  = f"=ROUND({N}{r}*{J}$2,2)"

        # K - DESPESA = CUSTO REAL × DESP%
        ws.cell(r, COL['DESP']).value   = f"=ROUND({N}{r}*{K}$2,2)"

        # L - CRED ICMS = NF_U × |CRED%|  (0 se ST/CST isento — ANT não zera o crédito)
        cst_isentos = f'{Oc}{r}="40",{Oc}{r}="60",{Oc}{r}="102",{Oc}{r}="500"'
        ws.cell(r, COL['CRED']).value   = (
            f"=IF(OR({G}{r}>0,{cst_isentos}),"
            f"0,ROUND({F}{r}*{Lc}$2,2))"
        )

        # M - CUSTO ENTRADA = C_REAL + ST + ANT + IPI + FRETE + DESP − CRED
        ws.cell(r, COL['C_ENT']).value  = (
            f"=ROUND({N}{r}+{G}{r}+{H}{r}+{I}{r}+{J}{r}+{K}{r}+{Lc}{r},2)"
        )

        # P - FEDERAL = P_VAR × FED%
        ws.cell(r, COL['FED']).value    = f"=ROUND({W}{r}*{P2}$2,2)"

        # Q - CARTÃO = P_VAR × CART%
        ws.cell(r, COL['CARTAO']).value = f"=ROUND({W}{r}*{Qc}$2,2)"

        # R - ICMS SAÍDA = P_VAR × ICMS%
        ws.cell(r, COL['ICMS_S']).value = f"=ROUND({W}{r}*{Rc}$2,2)"

        # S - CUSTO SAÍDA = C_ENT + FED + CART + ICMS
        ws.cell(r, COL['C_SAIDA']).value = (
            f"=ROUND({M}{r}+{P2}{r}+{Qc}{r}+{Rc}{r},2)"
        )

        # U - PREÇO MÍN VIÁVEL = C_SAIDA ÷ (1 − META%)
        ws.cell(r, COL['P_MIN']).value  = (
            f"=IF({T}{r}>0,ROUND({S}{r}/(1-{T}{r}),2),0)"
        )

        # W - PREÇO VAREJO: usa PREÇO ATUAL se > 0, senão arredondar_99(C_REAL × 2)
        ws.cell(r, COL['P_VAR']).value  = (
            f"=IF({V}{r}>0,{V}{r},"
            f"IF({N}{r}<=0,0,INT({N}{r}*2)-IF({N}{r}*2-INT({N}{r}*2)<=0.5,1,0)+0.99))"
        )

        # X - MARGEM REAL = (P_VAR − C_SAIDA) / P_VAR
        ws.cell(r, COL['MARGEM']).value = (
            f"=IF({W}{r}>0,ROUND(({W}{r}-{S}{r})/{W}{r},4),0)"
        )

        # ── Atacado ───────────────────────────────────────────────────────────

        # Y - NF ATC = NF_U × MULT_ATC
        ws.cell(r, COL['NF_ATC']).value = f"=ROUND({F}{r}*{Ya}$2,2)"

        # Z - PREÇO ATC = P_VAR × (1 − DESC_ATC%)
        ws.cell(r, COL['P_ATC']).value = f"=ROUND({W}{r}*(1-{Za}$2),2)"

        # AA - FEDERAL ATC = NF_ATC × FED%
        ws.cell(r, COL['FED_ATC']).value = f"=ROUND({Ya}{r}*{P2}$2,2)"

        # AB - CARTÃO ATC = P_ATC × CART%
        ws.cell(r, COL['CART_ATC']).value = f"=ROUND({Za}{r}*{Qc}$2,2)"

        # AC - ICMS ATC = NF_ATC × ICM%
        ws.cell(r, COL['ICM_ATC']).value = f"=ROUND({Ya}{r}*{Rc}$2,2)"

        # AD - CUSTO SAÍDA ATC = C_ENT + FED_ATC + CART_ATC + ICM_ATC
        ws.cell(r, COL['C_SAIDA_ATC']).value = (
            f"=ROUND({M}{r}+{AA}{r}+{AB}{r}+{AC}{r},2)"
        )

        # AE - MARGEM ATC = (P_ATC − C_SAIDA_ATC) / P_ATC
        ws.cell(r, COL['MARGEM_ATC']).value = (
            f"=IF({Za}{r}>0,ROUND(({Za}{r}-{AD}{r})/{Za}{r},4),0)"
        )

        # ── Formato numérico para colunas de fórmula ─────────────────────────
        for col_idx, fmt in [
            (COL['C_REAL'],      '#,##0.00'),
            (COL['FRETE'],       '#,##0.00'),
            (COL['DESP'],        '#,##0.00'),
            (COL['CRED'],        '#,##0.00'),
            (COL['C_ENT'],       '#,##0.00'),
            (COL['FED'],         '#,##0.00'),
            (COL['CARTAO'],      '#,##0.00'),
            (COL['ICMS_S'],      '#,##0.00'),
            (COL['C_SAIDA'],     '#,##0.00'),
            (COL['P_MIN'],       '#,##0.00'),
            (COL['P_VAR'],       '#,##0.00'),
            (COL['MARGEM'],      '0.00%'),
            (COL['NF_ATC'],      '#,##0.00'),
            (COL['P_ATC'],       '#,##0.00'),
            (COL['FED_ATC'],     '#,##0.00'),
            (COL['CART_ATC'],    '#,##0.00'),
            (COL['ICM_ATC'],     '#,##0.00'),
            (COL['C_SAIDA_ATC'], '#,##0.00'),
            (COL['MARGEM_ATC'],  '0.00%'),
        ]:
            ws.cell(r, col_idx).number_format = fmt

        # ── Borda e alinhamento em todas as células da linha ──────────────────
        for c in range(1, TOTAL_COLS + 1):
            cell = ws.cell(r, c)
            cell.border    = BORDA
            cell.alignment = ALI_CTR

        # ── Cores ─────────────────────────────────────────────────────────────
        _skip = {COL['AUDIT_SYS'], COL['AUDIT_EMB'], COL['META']}
        if has_st:
            # ST: peach em toda a linha, sem exceção
            for c in range(1, TOTAL_COLS + 1):
                if c not in _skip:
                    ws.cell(r, c).fill = F_PELE
        elif has_ant:
            # ANT: verde menta na linha inteira...
            for c in range(1, TOTAL_COLS + 1):
                if c not in _skip:
                    ws.cell(r, c).fill = F_AMAR_ANT
            # ...mas preserva as cores especiais de preço por cima
            ws.cell(r, COL['P_VAR']).fill      = F_AZUL
            ws.cell(r, COL['MARGEM']).fill     = F_AZUL
            ws.cell(r, COL['P_MIN']).fill      = F_AMAR
            ws.cell(r, COL['P_ATC']).fill      = F_AMAR
            ws.cell(r, COL['MARGEM_ATC']).fill = F_AMAR
        else:
            ws.cell(r, COL['P_VAR']).fill      = F_AZUL
            ws.cell(r, COL['MARGEM']).fill     = F_AZUL
            ws.cell(r, COL['P_MIN']).fill      = F_AMAR
            ws.cell(r, COL['P_ATC']).fill      = F_AMAR
            ws.cell(r, COL['MARGEM_ATC']).fill = F_AMAR

    # ── Largura das colunas ───────────────────────────────────────────────────
    widths = {
        1:10, 2:52, 3:14, 4:16, 5:7,
        6:10, 7:10, 8:10, 9:10,
        10:12, 11:10, 12:10, 13:12,
        14:8,  15:15,
        16:11, 17:10, 18:12, 19:12,
        20:9,  21:16,
        22:12, 23:14, 24:12,
        25:10, 26:12,
        27:12, 28:11, 29:10,
        30:15, 31:12,
        32:14, 33:9,
    }
    for c, w in widths.items():
        ws.column_dimensions[L(c)].width = w

    ws.freeze_panes = 'A3'
    wb.calculation.calcMode      = 'auto'
    wb.calculation.fullCalcOnLoad = True
    wb.save(path)


# ─── Dashboard HTML ───────────────────────────────────────────────────────────
def gerar_dashboard_html(rows_data, lucro_total=0.0, metricas=None):
    if not rows_data:
        return ""

    total     = len(rows_data)
    com_st    = sum(1 for r in rows_data if r['st_u']  > 0.005)
    com_ant   = sum(1 for r in rows_data if r['ant_u'] > 0.005 and r['st_u'] <= 0.005)
    sem_ambos = total - com_st - com_ant
    sem_preco = sum(1 for r in rows_data if r['p_atual'] <= 0)
    total_nf  = sum(r['nf_u'] * r['qtd'] for r in rows_data)

    margens = []
    for r in rows_data:
        if r['p_atual'] > 0 and r['nf_u'] > 0:
            margens.append((r['p_atual'] - r['nf_u']) / r['p_atual'])
    margem_media = (sum(margens) / len(margens) * 100) if margens else 0
    cor_margem   = "#E8F5E9" if margem_media >= 15 else "#FFEBEE"
    cor_lucro    = "#E2EFDA" if lucro_total >= 0 else "#FFEBEE"

    def fmt_brl(v):
        return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    def card(titulo, valor, cor, sub=""):
        return f"""<div style="background:{cor};border-radius:10px;padding:16px 20px;
                    flex:1;min-width:140px;box-shadow:0 2px 6px rgba(0,0,0,.07)">
          <div style="font-size:11px;color:#666;font-weight:600;text-transform:uppercase;
                      letter-spacing:.5px;margin-bottom:4px">{titulo}</div>
          <div style="font-size:26px;font-weight:700;color:#1a1a2e">{valor}</div>
          <div style="font-size:11px;color:#888;margin-top:2px">{sub}</div>
        </div>"""

    cards = "".join([
        card("Total Itens",     total,                 "#EBF5FB", "produtos na NF"),
        card("Com ST",          com_st,                "#FCE4D6", "subst. tributária"),
        card("Com ANT",         com_ant,               "#D1FAE5", "antecipação tributária"),
        card("Normal",          sem_ambos,             "#E2EFDA", "tributação normal"),
        card("Sem Preço Sist.", sem_preco,             "#FFF9C4", "usarão cálculo auto"),
        card("Valor Total NF",  fmt_brl(total_nf),     "#EBF5FB", "soma NF unit × qtd"),
        card("Margem Estimada", f"{margem_media:.1f}%", cor_margem, "preço atual vs NF unit"),
        card("Lucro Estimado",  fmt_brl(lucro_total),  cor_lucro, "meta 15% · preço varejo − custo saída"),
    ])

    pct_st    = total and (com_st    / total * 100)
    pct_ant   = total and (com_ant   / total * 100)
    pct_norm  = 100 - pct_st - pct_ant

    def _seg(w, bg, tc, label):
        if w < 1:
            return ''
        return (f'<div style="width:{w:.0f}%;background:{bg};display:flex;align-items:center;'
                f'justify-content:center;color:{tc};font-size:12px;font-weight:600">'
                f'{label} {w:.0f}%</div>')

    barra = f"""
    <div style="margin:20px 0 4px">
      <div style="font-size:11px;color:#666;font-weight:600;margin-bottom:6px;
                  text-transform:uppercase;letter-spacing:.5px">Distribuição da NF</div>
      <div style="display:flex;height:26px;border-radius:6px;overflow:hidden">
        {_seg(pct_norm, '#BDD7EE', '#1a5276', 'Normal')}
        {_seg(pct_st,   '#FCE4D6', '#784212', 'ST')}
        {_seg(pct_ant,  '#D1FAE5', '#065F46', 'ANT')}
      </div>
    </div>"""

    alertas_html = ""
    alertas = [r for r in rows_data if r['p_atual'] <= 0]
    if alertas:
        linhas = "".join(
            f"<tr><td style='padding:5px 10px'>{a['ref']}</td>"
            f"<td style='padding:5px 10px'>{a['desc'][:60]}</td>"
            f"<td style='padding:5px 10px;color:#c0392b;font-weight:600'>Sem preço no sistema</td></tr>"
            for a in alertas
        )
        alertas_html = f"""
        <div style="margin-top:20px">
          <div style="font-weight:700;font-size:13px;color:#c0392b;margin-bottom:8px">
            ⚠️ {len(alertas)} produto(s) sem preço no sistema — preço calculado será usado
          </div>
          <table style="width:100%;border-collapse:collapse;font-size:12px;background:#fdf2f2;
                        border-radius:8px;overflow:hidden">
            <thead><tr style="background:#fdecea">
              <th style="padding:6px 10px;text-align:left">REF</th>
              <th style="padding:6px 10px;text-align:left">DESCRIÇÃO</th>
              <th style="padding:6px 10px;text-align:left">STATUS</th>
            </tr></thead>
            <tbody>{linhas}</tbody>
          </table>
        </div>"""

    # ── Gráfico de lucro por produto ─────────────────────────────────────────
    chart_html = ""
    if metricas:
        produtos = sorted(
            [{'desc': r['desc'][:48], 'lucro': round(m['lucro'], 2)}
             for r, m in zip(rows_data, metricas)],
            key=lambda x: x['lucro'], reverse=True
        )
        if produtos:
            max_abs   = max(abs(p['lucro']) for p in produtos) or 1
            lucro_100 = sum(p['lucro'] for p in produtos)
            cjson     = json.dumps(produtos, ensure_ascii=False)

            bars = ""
            for i, p in enumerate(produtos):
                w   = abs(p['lucro']) / max_abs * 100
                cor = '#BDD7EE' if p['lucro'] >= 0 else '#FCE4D6'
                lc  = '#1a252f' if p['lucro'] >= 0 else '#c0392b'
                bars += (
                    f'<div style="display:flex;align-items:center;gap:10px">'
                    f'<span style="font-size:11px;color:#555;width:195px;min-width:195px;'
                    f'overflow:hidden;text-overflow:ellipsis;white-space:nowrap" '
                    f'title="{p["desc"]}">{p["desc"]}</span>'
                    f'<div style="flex:1;background:#f0f4f8;border-radius:4px;height:18px;overflow:hidden">'
                    f'<div id="ph-bar-{i}" style="height:100%;width:{w:.1f}%;background:{cor};'
                    f'border-radius:4px;transition:width 0.25s ease"></div></div>'
                    f'<span id="ph-lbl-{i}" style="font-size:11px;font-weight:700;color:{lc};'
                    f'min-width:100px;text-align:right">{fmt_brl(p["lucro"])}</span>'
                    f'</div>'
                )

            chart_html = f"""
      <div style="margin-top:24px;border-top:1px solid #e8e8e8;padding-top:20px">
        <div style="font-size:11px;color:#666;font-weight:600;text-transform:uppercase;
                    letter-spacing:.5px;margin-bottom:16px">Lucro Estimado por Produto</div>
        <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px">
          <span style="font-size:12px;color:#666;white-space:nowrap">% Vendido:</span>
          <input type="range" id="ph-slider" min="0" max="100" value="100"
                 style="flex:1;accent-color:#3b82f6;cursor:pointer"
                 oninput="phUpdate(this.value)">
          <span id="ph-pct" style="font-size:14px;font-weight:700;color:#1a1a2e;
                                    min-width:42px;text-align:right">100%</span>
        </div>
        <div style="background:#EBF5FB;border-radius:8px;padding:10px 16px;margin-bottom:14px;
                    display:flex;justify-content:space-between;align-items:center">
          <span style="font-size:11px;color:#666;font-weight:600;text-transform:uppercase;
                       letter-spacing:.5px">Total Projetado</span>
          <span id="ph-total" style="font-size:20px;font-weight:700;color:#1a252f">{fmt_brl(lucro_100)}</span>
        </div>
        <div style="max-height:380px;overflow-y:auto;padding-right:6px;
                    display:flex;flex-direction:column;gap:7px">
          {bars}
        </div>
      </div>
      <script>
      (function(){{
        var D={cjson}, MAX={max_abs};
        function brl(v){{
          var n=v<0, s=Math.abs(v).toFixed(2).replace('.',',')
            .replace(/\B(?=(\d{{3}})+(?!\d))/g,'.');
          return (n?'-':'')+'R$ '+s;
        }}
        window.phUpdate=function(pct){{
          var f=pct/100;
          document.getElementById('ph-pct').textContent=pct+'%';
          var tot=0;
          D.forEach(function(p,i){{
            var v=p.lucro*f; tot+=v;
            var b=document.getElementById('ph-bar-'+i);
            var l=document.getElementById('ph-lbl-'+i);
            if(b) b.style.width=(Math.abs(p.lucro)/MAX*100*f)+'%';
            if(l){{l.textContent=brl(v); l.style.color=v<0?'#c0392b':'#1a252f';}}
          }});
          var t=document.getElementById('ph-total');
          if(t){{t.textContent=brl(tot); t.style.color=tot<0?'#c0392b':'#1a252f';}}
        }};
      }})();
      </script>"""

    return f"""
    <div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
                margin-top:32px;padding:24px;background:#fff;
                border-radius:14px;border:1px solid #e0e0e0;
                box-shadow:0 2px 12px rgba(0,0,0,.06)">
      <h3 style="margin:0 0 18px;font-size:17px;color:#1a252f;
                 border-bottom:2px solid #BDD7EE;padding-bottom:10px">
        📊 Dashboard da NF
      </h3>
      <div style="display:flex;gap:12px;flex-wrap:wrap">{cards}</div>
      {barra}
      {alertas_html}
      {chart_html}
    </div>"""
